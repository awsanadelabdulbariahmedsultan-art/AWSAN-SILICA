# ====================================================================
# 🌍 AWSAN SILICA PROJECT - ADVANCED EXCEL WORKBOOK GENERATOR
# Powered by: Eng. Awsan Adel Abdulbari Ahmed Sultan
# Owned by: Awsan Dew For Marketing Services (Yemen)
# ====================================================================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_advanced_revenue_model():
    # 1. Initialize Workbook and Sheets
    wb = Workbook()
    
    # ----------------------------------------------------------------
    # SHEET 1: Revenue Breakdown By Tier
    # ----------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Revenue Breakdown By Tier"
    ws1.views.sheetView[0].showGridLines = True
    
    # Header Data
    ws1["A1"] = "AWSAN SILICA PROJECT - REVENUE STREAM MATRIX (2026)"
    ws1["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    ws1["A1"].fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    ws1.merge_cells("A1:E1")
    ws1["A1"].alignment = Alignment(horizontal="center")
    
    # Metadata rows
    ws1["A2"] = "Proprietary Owner: Awsan Dew For Marketing Services"
    ws1["A2"].font = Font(name="Segoe UI", size=10, italic=True)
    ws1["A3"] = "Chairman & Chief Systems Architect: Eng. Awsan Adel Abdulbari Ahmed Sultan"
    ws1["A3"].font = Font(name="Segoe UI", size=10, italic=True)
    
    # Table Headings
    headers1 = [
        "Product Tier / مستوى المنتج النهائي", 
        "Target Industry / القطاع المستهدف", 
        "Annual Yield (Tons) / الإنتاج (طن)", 
        "Price per Ton (USD) / السعر ($)", 
        "Projected Revenue / العائد السنوي المتوقع ($)"
    ]
    
    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=5, column=col_num)
        cell.value = header
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2A4D69", end_color="2A4D69", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Table Data (Rows 6 to 9)
    data1 = [
        ["Tier 1: Raw White Sand / الرمل الأبيض الخام", "Construction & Basic Foundry / البناء والسباكة", 200000, 20],
        ["Tier 2: Refined Micro-Powder / مسحوق السيليكا المكرر", "Premium Optics & Luxury Glass / البصريات والزجاج", 175000, 250],
        ["Tier 3: High-Purity Quartz (HPQ) / الكوارتز فائق النقاء", "Solar PV & Semiconductors / الألواح والرقائق", 75000, 4000],
        ["Tier 4: Polished BretonStone Slabs / ألواح الكوارتز الفاخرة", "Luxury Architecture & Cleanrooms / الأسطح الفاخرة والطبية", 50000, 3500]
    ]
    
    for r_idx, row_data in enumerate(data1, 6):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Segoe UI", size=11)
            if c_idx in:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
        # Apply formula for total row revenue
        res_cell = ws1.cell(row=r_idx, column=5)
        res_cell.value = f"=C{r_idx}*D{r_idx}"
        res_cell.font = Font(name="Segoe UI", size=11, bold=True)
        res_cell.number_format = '$#,##0'
        res_cell.alignment = Alignment(horizontal="right")

    # Totals Row (Row 10)
    ws1["A10"] = "TOTAL GROSS OUTPUT & REVENUE"
    ws1["A10"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws1["A10"].fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    ws1.merge_cells("A10:B10")
    
    ws1["C10"] = "=SUM(C6:C9)"
    ws1["C10"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws1["C10"].fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    ws1["C10"].number_format = '#,##0'
    
    ws1["D10"] = "" # Empty for aesthetic alignment
    ws1["D10"].fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    
    ws1["E10"] = "=SUM(E6:E9)"
    ws1["E10"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws1["E10"].fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    ws1["E10"].number_format = '$#,##0'
    ws1["E10"].alignment = Alignment(horizontal="right")

    # Adjust Column Widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ----------------------------------------------------------------
    # SHEET 2: ROI & Financial Viability
    # ----------------------------------------------------------------
    ws2 = wb.create_sheet(title="ROI & Financial Viability")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "AWSAN SILICA - CAPITAL EFFICIENCY & VALUATION MODELS"
    ws2["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    ws2["A1"].fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    ws2.merge_cells("A1:C1")
    ws2["A1"].alignment = Alignment(horizontal="center")
    
    ws2["A3"] = "Financial Metric / مؤشر الكفاءة المالية"
    ws2["B3"] = "Excel Reference / المعادلات والربط"
    ws2["C3"] = "Computed Value / القيمة المالية المحسوبة"
    
    for col_num in range(1, 4):
        cell = ws2.cell(row=3, column=col_num)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2A4D69", end_color="2A4D69", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Core Metric Rows
    metrics = [
        ("Total Capital Expenditure (CAPEX) / إجمالي الاستثمار الرأسمالي", "Fixed Baseline Container", 125000000, "$#,##0"),
        ("Annual Operational Expenditure (OPEX) / إجمالي النفقات التشغيلية", "Fixed Overhead Container", 45000000, "$#,##0"),
        ("Projected Gross Annual Revenue / إجمالي الإيرادات السنوية الإجمالية", "='Revenue Breakdown By Tier'!E10", "='Revenue Breakdown By Tier'!E10", "$#,##0"),
        ("💰 NET ANNUAL CASH FLOW (NET PROFIT) / صافي الأرباح السنوية المباشرة", "=C6-C5", "=C6-C5", "$#,##0"),
        ("Return on Investment (ROI) Efficiency / معدل العائد الاستثماري السنوي", "=(C7/C4)*100", "=(C7/C4)", "0.00%"),
        ("Capital Payback Period (Months) / فترة استرداد رأس المال بالشهور", "=(C4/C7)*12", "=(C4/C7)*12", "0.0")
    ]
    
    for idx, (metric, ref, formula, num_fmt) in enumerate(metrics, 4):
        ws2.cell(row=idx, column=1, value=metric).font = Font(name="Segoe UI", size=11)
        ws2.cell(row=idx, column=2, value=ref).font = Font(name="Segoe UI", size=10, italic=True, color="7F8C8D")
        
        val_cell = ws2.cell(row=idx, column=3, value=formula)
        val_cell.font = Font(name="Segoe UI", size=11, bold=True)
        val_cell.number_format = num_fmt
        val_cell.alignment = Alignment(horizontal="right")
        
        # Highlight important rows (Profit and ROI)
        if idx in:
            ws2.cell(row=idx, column=1).font = Font(name="Segoe UI", size=11, bold=True, color="004D40")
            val_cell.fill = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
            
    # Add System Footer Tag
    ws2["A11"] = "System Intellectual Status: Secure Portfolio / Powered by Eng. Awsan Adel"
    ws2["A11"].font = Font(name="Segoe UI", size=9, italic=True, color="7F8C8D")
    ws2.merge_cells("A11:C11")

    # Adjust Widths for Sheet 2
    ws2.column_dimensions["A"].width = 65
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 25

    # Save output
    filename = "revenue_projection_model.xlsx"
    wb.save(filename)
    print("====================================================================")
    print(f"✅ SUCCESS: Multi-sheet financial model '{filename}' built successfully!")
    print("• Sheet 1: 'Revenue Breakdown By Tier' with mathematical SUM and multiplication logics.")
    print("• Sheet 2: 'ROI & Financial Viability' with cell cross-referencing formulas.")
    print("• Encryption Status: Secure asset for Awsan Dew For Marketing Services.")
    print("====================================================================")

if __name__ == "__main__":
    create_advanced_revenue_model()

